"""
Pre-Upgrade Checklist Workflow

Workflow for filling the Pre-Upgrade Checklist Excel template.
Writes approved cell values (from generate_upgrade_readiness_report) into an Excel template copy.

Produces two sheets: Summary (built dynamically) and Pre-Upgrade Checklist.

The readiness report caches the full cell_values payload on disk keyed by the
CMC cluster name. write_checklist_excel prefers the cached blob over the
cell_values_json argument so the LLM does not need to ferry the (often large)
JSON between tool calls — any truncation / paraphrasing by the model is
bypassed entirely.
"""

import base64
import json
import os
import shutil
import sys
import tempfile
import time
from pathlib import Path
from typing import Optional, TypedDict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ---------------------------------------------------------------------------
# Server-side cache for checklist data (keyed by CMC cluster name)
# ---------------------------------------------------------------------------
# The cache lives under /app/data/checklist_cache/ inside the container,
# alongside /app/data/tokens/ (Cloud Portal JWTs). Entries expire after 24h
# so a stale report can't silently overwrite a fresh run.

CACHE_DIR = Path(os.getenv("CHECKLIST_CACHE_DIR", "/app/data/checklist_cache"))
CACHE_TTL_SECONDS = 24 * 60 * 60  # 24 hours


def _safe_cluster_key(cmc_cluster_name: str) -> str:
    """Sanitize a cluster name for use as a filename (keep it simple)."""
    import re
    return re.sub(r"[^A-Za-z0-9._-]", "_", (cmc_cluster_name or "").strip())[:128] or "default"


def save_checklist_cache(cmc_cluster_name: str, payload: dict) -> Optional[Path]:
    """Persist the full checklist payload so write_checklist_excel can read it later.

    `payload` is the same dict that would otherwise be JSON-serialised into
    the <checklist_data> block of the readiness report. Keeps `_summary` and
    all per-row details intact.

    Returns the path written (or None on write failure — cache is best-effort).
    """
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        key = _safe_cluster_key(cmc_cluster_name)
        path = CACHE_DIR / f"{key}.json"
        tmp = path.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(payload))
        tmp.replace(path)
        return path
    except OSError:
        return None


def load_checklist_cache(cmc_cluster_name: str) -> Optional[dict]:
    """Load a previously cached payload. Returns None if missing or stale."""
    try:
        key = _safe_cluster_key(cmc_cluster_name)
        path = CACHE_DIR / f"{key}.json"
        if not path.exists():
            return None
        age = time.time() - path.stat().st_mtime
        if age > CACHE_TTL_SECONDS:
            return None
        return json.loads(path.read_text())
    except (OSError, json.JSONDecodeError):
        return None


# ---------------------------------------------------------------------------
# State
# ---------------------------------------------------------------------------

class ChecklistState(TypedDict):
    cmc_cluster_name: str
    cloud_cluster_name: str
    from_version: str
    to_version: str
    template_path: str
    output_path: str
    cluster_metadata: dict
    validation_checks: dict
    cloud_metadata: dict
    upgrade_knowledge: list
    cell_values: dict
    errors: list
    report: str


# ---------------------------------------------------------------------------
# Map collected data to cell values (used internally by readiness_report.py)
# ---------------------------------------------------------------------------

def map_data_to_cells(state: ChecklistState) -> ChecklistState:
    """Transform collected data into Excel cell values for the 2-sheet checklist.

    Row indices map to rows in templates/pre_upgrade_checklist.xlsx (see the
    build_template.py companion script that generated that file). The writer
    places dict key "B" into col C (Details / Findings) and "C" into col D
    (Status); col B of the template is a pre-filled label.
    """
    meta = state.get("cluster_metadata", {})
    checks = state.get("validation_checks", {})
    cloud = state.get("cloud_metadata", {})
    knowledge = state.get("upgrade_knowledge", [])
    jira_issues = state.get("jira_issues", {})
    zendesk_issues = state.get("zendesk_issues", {})

    cells = {}
    is_cloud = meta.get("deployment_type", {}).get("is_cloud", False)

    def _tag(value, source):
        """Append [Source: ...] to the value so the Excel sheet shows data provenance."""
        return f"{value}\n[Source: {source}]"

    def _status_from_check(check_status: str) -> str:
        """Map validation-check status (PASS/WARNING/FAIL) to checklist status."""
        return {"PASS": "Done", "WARNING": "Review", "FAIL": "Action Required"}.get(
            check_status, "Review"
        )

    integrations = meta.get("integrations", {})
    integration_items = integrations.get("integrations", {}) if isinstance(integrations, dict) else {}

    # =========================================================================
    # Section 1: ENVIRONMENT OVERVIEW  (rows 12-19)
    # =========================================================================

    # --- R12 · 1.0 Upgrade Path ---
    dep = meta.get("deployment_type", {})
    cells[12] = {
        "attribute": "Upgrade Path",
        "B": _tag(
            f"From: {state.get('from_version', 'N/A')}\nTo: {state.get('to_version', 'N/A')}",
            "CMC",
        ),
        "C": "Done",
    }

    # --- R13 · 2.0 Deployment ---
    dep_value = f"{dep.get('deployment_type', 'Unknown')} ({dep.get('cloud_provider', 'Unknown')})"
    cells[13] = {"attribute": "Deployment", "B": _tag(dep_value, "CMC"), "C": "Done"}

    # --- R14 · 3.0 Topology ---
    topo = meta.get("topology", {})
    topo_value = f"{topo.get('topology_type', 'Unknown')} ({topo.get('node_count', '?')} nodes)"
    if topo.get("is_ha"):
        topo_value += " - HA Enabled"
    nodes = topo.get("nodes", [])
    if nodes:
        node_lines = [
            f"  {n.get('name', '?')}: {n.get('type', '?')} - Services: {', '.join(n.get('services', []))}"
            for n in nodes
        ]
        topo_value += "\n" + "\n".join(node_lines)
    cells[14] = {"attribute": "Topology", "B": _tag(topo_value, "CMC"), "C": "Done"}

    # --- R15 · 4.0 Database ---
    db = meta.get("database", {})
    if db.get("migration_needed"):
        cells[15] = {
            "attribute": "Database",
            "B": _tag(f"{db.get('db_type', 'Oracle')} — migration to MySQL needed", "CMC"),
            "C": "Action Required",
        }
    else:
        cells[15] = {
            "attribute": "Database",
            "B": _tag(f"{db.get('db_type', 'Unknown')} — no migration needed", "CMC"),
            "C": "Done",
        }

    # --- R16 · 5.0 Spark Version ---
    spark_from_cloud = cloud.get("spark_version")
    spark_ver = spark_from_cloud or meta.get("infrastructure", {}).get("spark_mode", "N/A")
    spark_source = "Cloud Portal" if spark_from_cloud else "CMC"
    cells[16] = {"attribute": "Spark Version", "B": _tag(str(spark_ver), spark_source), "C": "Done"}

    # --- R17 · 6.0 Python Version ---
    python_ver = cloud.get("python_version", "N/A")
    cells[17] = {"attribute": "Python Version", "B": _tag(str(python_ver), "Cloud Portal"), "C": "Done"}

    # --- R18 · 7.0 Connectors (overview) ---
    features = meta.get("features", {})
    enabled_connectors = features.get("connectors", [])
    disabled_connectors = features.get("disabled_connectors", [])
    conn_lines = []
    if enabled_connectors:
        conn_lines.append(f"Enabled ({len(enabled_connectors)}): {', '.join(enabled_connectors)}")
    if disabled_connectors:
        conn_lines.append(f"Disabled ({len(disabled_connectors)}): {', '.join(disabled_connectors)}")
    if not conn_lines:
        conn_lines.append("No connectors found")
    cells[18] = {
        "attribute": "Connectors",
        "B": _tag("\n".join(conn_lines), "CMC"),
        "C": "Done",
    }

    # --- R19 · 8.0 Tenants (overview, with MS Synced) ---
    ts = meta.get("tenant_storage", {})
    if ts.get("status") == "success":
        tenant_lines = []
        for t in ts.get("tenants", []):
            name = t.get("name", "?")
            quota = t.get("disk_quota", "unknown")
            unit = t.get("disk_unit", "")
            enabled = "Enabled" if t.get("enabled") else "Disabled"
            # Fix #6: surface MS Synced per-tenant
            ms = "MS Synced" if t.get("isMSSynced") or t.get("ms_synced") else "Not MS Synced"
            tenant_lines.append(f"{name}: {quota} {unit} ({enabled}, {ms})")
        cells[19] = {
            "attribute": "Tenants",
            "B": _tag("\n".join(tenant_lines) if tenant_lines else "No tenants", "CMC"),
            "C": "Done",
        }
    else:
        cells[19] = {
            "attribute": "Tenants",
            "B": _tag(f"Could not retrieve: {ts.get('error', 'unknown')}", "CMC"),
            "C": "Failed",
        }

    # =========================================================================
    # Section 2: RELEASE NOTES & UPGRADE CONSIDERATIONS  (rows 21-23)
    # =========================================================================

    # --- R21 · 9.0 Data Agent (from cloud_metadata if available) ---
    if cloud:
        da_enabled = cloud.get("data_agent_enabled", False)
        cells[21] = {
            "attribute": "Data Agent",
            "B": _tag(
                f"Data Agent: {'Enabled' if da_enabled else 'Disabled'}"
                + (" — confirm DA upgrade with customer" if da_enabled else ""),
                "Cloud Portal",
            ),
            "C": "Review" if da_enabled else "Done",
        }
    else:
        cells[21] = {
            "attribute": "Data Agent",
            "B": _tag("N/A (cloud data not available)", "Cloud Portal"),
            "C": "N/A",
        }

    # --- R22 · 10.0 Behavior Changes (from knowledge base) ---
    if knowledge:
        items = [k.get("title", k.get("text", "N/A"))[:120] for k in knowledge[:5]]
        cells[22] = {
            "attribute": "Behavior Changes",
            "B": _tag("\n".join(f"- {item}" for item in items), "Knowledge Base"),
            "C": "Review",
        }
    else:
        cells[22] = {
            "attribute": "Behavior Changes",
            "B": _tag("No behavior-change notes found for this version pair", "Knowledge Base"),
            "C": "Review",
        }

    # --- R23 · 11.0 Previous Considerations (placeholder, human-filled) ---
    cells[23] = {
        "attribute": "Previous Considerations",
        "B": "Upgrade considerations from prior versions may also apply unless stated otherwise — review release notes.",
        "C": "Review",
    }

    # =========================================================================
    # Section 3: AUTOMATED HEALTH CHECKS  (rows 25-34)
    # =========================================================================

    # --- R25 · 12.0 Service Status ---
    svc_check = checks.get("Service Status", {})
    cells[25] = {
        "attribute": "Service Status",
        "B": _tag("\n".join(svc_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(svc_check.get("status", "N/A")),
    }

    # --- R26 · 13.0 Memory Status ---
    mem_check = checks.get("Memory Status", {})
    mem_details = mem_check.get("details", [])
    cells[26] = {
        "attribute": "Memory Status",
        "B": _tag("\n".join(mem_details) or "N/A", "CMC"),
        "C": _status_from_check(mem_check.get("status", "N/A")),
    }

    # --- R27 · 14.0 Cluster Configuration (auto-start suppression on cloud lives in the check itself) ---
    cfg_check = checks.get("Cluster Configuration", {})
    cells[27] = {
        "attribute": "Cluster Configuration",
        "B": _tag("\n".join(cfg_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(cfg_check.get("status", "N/A")),
    }

    # --- R28 · 15.0 Infrastructure Services ---
    infra_check = checks.get("Infrastructure Services", {})
    cells[28] = {
        "attribute": "Infrastructure Services",
        "B": _tag("\n".join(infra_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(infra_check.get("status", "N/A")),
    }

    # --- R29 · 16.0 Node Topology (same-service HA logic already in the check) ---
    nt_check = checks.get("Node Topology", {})
    cells[29] = {
        "attribute": "Node Topology",
        "B": _tag("\n".join(nt_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(nt_check.get("status", "N/A")),
    }

    # --- R30 · 17.0 Connectors (health / test connection) ---
    conn_check = checks.get("Connectors", {})
    cells[30] = {
        "attribute": "Connectors",
        "B": _tag("\n".join(conn_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(conn_check.get("status", "N/A")),
    }

    # --- R31 · 18.0 Tenants (health) ---
    tenants_check = checks.get("Tenants", {})
    cells[31] = {
        "attribute": "Tenants",
        "B": _tag("\n".join(tenants_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(tenants_check.get("status", "N/A")),
    }

    # --- R32 · 19.0 Email / SMTP  [Fix #2] ---
    email_check = checks.get("Email Configuration", {})
    cells[32] = {
        "attribute": "Email / SMTP",
        "B": _tag("\n".join(email_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(email_check.get("status", "N/A")),
    }

    # --- R33 · 20.0 Notebook & SQLi  [Fix #3] ---
    nb_check = checks.get("Notebook & SQLi", {})
    cells[33] = {
        "attribute": "Notebook & SQLi",
        "B": _tag("\n".join(nb_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(nb_check.get("status", "N/A")),
    }

    # --- R34 · 21.0 Database Migration ---
    dbm_check = checks.get("Database Migration", {})
    cells[34] = {
        "attribute": "Database Migration",
        "B": _tag("\n".join(dbm_check.get("details", [])) or "N/A", "CMC"),
        "C": _status_from_check(dbm_check.get("status", "N/A")),
    }

    # =========================================================================
    # Section 4: STORAGE & RESOURCES  (rows 36-39)
    # =========================================================================

    # --- R36 · 22.0 Disk Space ---
    if cloud:
        disk_lines = [
            f"Analytics Pod: {cloud.get('data_size_gb', '?')} GB (allocated)",
            f"Loader Pod: {cloud.get('loader_size_gb', '?')} GB (allocated)",
            f"CMC Pod: {cloud.get('cmc_size_gb', '?')} GB (allocated)",
            "Note: Per-pod utilization not available (API limited)",
            f"Tenant Folder Size: {cloud.get('consumed_data_gb', '?')} GB, "
            f"Available Disk: {cloud.get('available_disk_gb', '?')} GB",
        ]
        cells[36] = {
            "attribute": "Disk Space",
            "B": _tag("\n".join(disk_lines), "Cloud Portal"),
            "C": "Done",
        }
    else:
        cells[36] = {
            "attribute": "Disk Space",
            "B": _tag("Cloud data not available", "Cloud Portal"),
            "C": "N/A",
        }

    # --- R37 · 23.0 Sizing ---
    if cloud:
        sizing_lines = []
        if cloud.get("analytics_size"):
            sizing_lines.append(f"Analytics: {cloud.get('analytics_size')}")
        if cloud.get("loader_size"):
            sizing_lines.append(f"Loader: {cloud.get('loader_size')}")
        cells[37] = {
            "attribute": "Sizing",
            "B": _tag("\n".join(sizing_lines) if sizing_lines else "Sizing details not available", "Cloud Portal"),
            "C": "Done" if sizing_lines else "Review",
        }
    else:
        cells[37] = {
            "attribute": "Sizing",
            "B": _tag("Cloud data not available", "Cloud Portal"),
            "C": "N/A",
        }

    # --- R38 · 24.0 Timezone ---
    cp_timezone = cloud.get("timezone") if cloud else None
    if cp_timezone:
        cells[38] = {
            "attribute": "Timezone",
            "B": _tag(f"Cluster Timezone: {cp_timezone}", "Cloud Portal"),
            "C": "Done",
        }
    else:
        tz_keys = [k for k in integration_items if "timezone" in k.lower() or "tz" in k.lower()]
        if tz_keys:
            tz_values = [f"{k}: enabled={integration_items[k].get('enabled', False)}" for k in tz_keys]
            cells[38] = {"attribute": "Timezone", "B": _tag("\n".join(tz_values), "CMC"), "C": "Done"}
        else:
            cells[38] = {
                "attribute": "Timezone",
                "B": _tag("No timezone config found (may use default)", "CMC"),
                "C": "Review",
            }

    # --- R39 · Auto-Suspend / Idle Time  [Fix #1 label: was "Environment Type"] ---
    if cloud:
        sleeppable = cloud.get("sleeppable")
        idle_hours = cloud.get("idle_time_hours")
        if sleeppable is not None:
            suspend_status = "Enabled" if sleeppable else "Disabled"
            idle_str = f"{idle_hours} hours" if idle_hours is not None else "N/A"
            cells[39] = {
                "attribute": "Auto-Suspend / Idle Time",
                "B": _tag(f"Auto-Suspend: {suspend_status}\nIdle Time: {idle_str}", "Cloud Portal"),
                "C": "Done",
            }
        else:
            cells[39] = {
                "attribute": "Auto-Suspend / Idle Time",
                "B": _tag("Auto-suspend config not available", "Cloud Portal"),
                "C": "Review",
            }
    else:
        cells[39] = {
            "attribute": "Auto-Suspend / Idle Time",
            "B": _tag("N/A (on-prem deployment)", "Cloud Portal"),
            "C": "N/A",
        }

    # =========================================================================
    # Section 5: ZENDESK TICKET ANALYSIS  (rows 41-42)
    # =========================================================================

    complete = zendesk_issues.get("complete_issues", {})
    zd_tickets = complete.get("issues", [])
    solved_tickets = [t for t in zd_tickets if str(t.get("status", "")).lower() in ("solved", "closed")]
    open_tickets = [t for t in zd_tickets if t not in solved_tickets]

    # --- R41 · 25.0 Solved Upgrade Tickets ---
    if solved_tickets:
        lines = []
        for t in solved_tickets[:8]:
            tid = t.get("ticket_id", "?")
            subj = str(t.get("subject", "?"))[:80]
            lines.append(f"#{tid} {subj} — SOLVED")
        if len(solved_tickets) > 8:
            lines.append(f"... and {len(solved_tickets) - 8} more")
        cells[41] = {
            "attribute": "Solved Upgrade Tickets",
            "B": _tag("\n".join(lines), "Zendesk"),
            "C": "Done",
        }
    else:
        cells[41] = {
            "attribute": "Solved Upgrade Tickets",
            "B": _tag("No solved tickets found for this upgrade path", "Zendesk"),
            "C": "Done",
        }

    # --- R42 · 26.0 Open Tickets on target version ---
    if open_tickets:
        ticket_lines = []
        for t in open_tickets[:10]:
            tid = t.get("ticket_id", "?")
            subj = str(t.get("subject", "?"))[:60]
            status = t.get("status", "?")
            fix = t.get("fixed_in", "")
            workaround = "Yes" if t.get("has_workaround") else "No"
            line = f"#{tid} [{status}] {subj} | Workaround: {workaround}"
            if fix:
                line += f" | Fix: {fix}"
            ticket_lines.append(line)
        if len(open_tickets) > 10:
            ticket_lines.append(f"... and {len(open_tickets) - 10} more")
        ticket_lines.append("")
        ticket_lines.append("Note: See Claude chat for detailed analysis and workarounds")
        cells[42] = {
            "attribute": "Open Tickets on target version",
            "B": _tag("\n".join(ticket_lines), "Zendesk"),
            "C": "Review",
        }
    else:
        cells[42] = {
            "attribute": "Open Tickets on target version",
            "B": _tag(
                "No open Zendesk tickets found for this upgrade path\n\n"
                "Note: See Claude chat for detailed analysis",
                "Zendesk",
            ),
            "C": "Done",
        }

    # =========================================================================
    # Section 6: JIRA BUG ANALYSIS  (rows 44-45)  [Fix #4 label]
    # =========================================================================

    jira_bugs = jira_issues.get("bugs", [])
    fixed_bugs = [b for b in jira_bugs if b.get("category") == "fixed_in_target"]
    still_open_bugs = [b for b in jira_bugs if b.get("category") == "still_open"]
    later_bugs = [b for b in jira_bugs if b.get("category") == "requires_later_release"]

    # --- R44 · 27.0 Bugs Fixed by Upgrade ---
    if fixed_bugs:
        lines = []
        for b in fixed_bugs[:8]:
            lines.append(
                f"{b.get('key', '?')} — {str(b.get('summary', '?'))[:70]} "
                f"(fix: {b.get('fix_version', '?')})"
            )
        if len(fixed_bugs) > 8:
            lines.append(f"... and {len(fixed_bugs) - 8} more")
        cells[44] = {
            "attribute": "Bugs Fixed by Upgrade",
            "B": _tag("\n".join(lines), "Jira"),
            "C": "Done",
        }
    else:
        cells[44] = {
            "attribute": "Bugs Fixed by Upgrade",
            "B": _tag("No bugs known to be fixed in this upgrade path", "Jira"),
            "C": "Done",
        }

    # --- R45 · 28.0 Open Bugs on target version  [Fix #4: label] ---
    if still_open_bugs or later_bugs:
        bug_lines = []
        if still_open_bugs:
            bug_lines.append(f"Still open ({len(still_open_bugs)}):")
            for b in still_open_bugs[:5]:
                bug_lines.append(
                    f"  {b.get('key', '?')} — {str(b.get('summary', '?'))[:60]}"
                )
        if later_bugs:
            bug_lines.append(f"Requires later release ({len(later_bugs)}):")
            for b in later_bugs[:5]:
                bug_lines.append(
                    f"  {b.get('key', '?')}, Fix Version: {b.get('fix_version', '?')}"
                )
        cells[45] = {
            "attribute": "Open Bugs on target version",
            "B": _tag("\n".join(bug_lines), "Jira"),
            "C": "Review",
        }
    else:
        cells[45] = {
            "attribute": "Open Bugs on target version",
            "B": _tag("No open bugs found for this upgrade path", "Jira"),
            "C": "Done",
        }

    # =========================================================================
    # Section 7: MANUAL PRE-UPGRADE CHECKS  (rows 47-54)
    # =========================================================================

    # --- R47 · 29.0 Scheduled Jobs ---
    cells[47] = {
        "attribute": "Scheduled Jobs",
        "B": "Pause scheduled jobs and SendNow dashboards before upgrade (manual step).",
        "C": "Pending",
    }

    # --- R48 · 30.0 Scheduled Load Jobs ---
    cells[48] = {
        "attribute": "Scheduled Load Jobs",
        "B": "Pause scheduled load jobs before upgrade (manual step).",
        "C": "Pending",
    }

    # --- R49 · 31.0 Run Alias Sync ---
    if is_cloud:
        cells[49] = {
            "attribute": "Run Alias Sync",
            "B": _tag("N/A (Cloud deployment) — Alias Sync is only relevant for on-prem installations.", "CMC"),
            "C": "N/A",
        }
    else:
        cells[49] = {
            "attribute": "Run Alias Sync",
            "B": "Run Alias Sync before upgrade (on-prem only).",
            "C": "Pending",
        }

    # --- R50 · 32.0 Run Inspector Tool  [Fix #5: deferred with reason] ---
    cells[50] = {
        "attribute": "Run Inspector Tool",
        "B": (
            "Deferred: No API available for Inspector Tool / Validation Dashboard. "
            "Upgrade engineer should run manually from CMC UI and confirm no Sev-1 "
            "issues before proceeding."
        ),
        "C": "N/A",
    }

    # --- R51 · 33.0 Download Incorta Package ---
    if is_cloud:
        cells[51] = {
            "attribute": "Download Incorta Package",
            "B": _tag("N/A (Cloud deployment) — package management is handled by Incorta Cloud.", "CMC"),
            "C": "N/A",
        }
    else:
        cells[51] = {
            "attribute": "Download Incorta Package",
            "B": "Download Incorta package to all nodes (on-prem only).",
            "C": "Pending",
        }

    # --- R52 · 34.0 Install Chromium ---
    if is_cloud:
        cells[52] = {
            "attribute": "Install Chromium",
            "B": _tag("N/A (Cloud deployment) — Chromium is pre-installed in container images.", "CMC"),
            "C": "N/A",
        }
    else:
        cells[52] = {
            "attribute": "Install Chromium",
            "B": "Install Chromium Headless Browser if needed for PDF exports (on-prem only).",
            "C": "Pending",
        }

    # --- R53 · 35.0 Rollback Validation ---
    cells[53] = {
        "attribute": "Rollback Validation",
        "B": "Perform a successful rollback test on the pre-prod environment before proceeding.",
        "C": "Pending",
    }

    # --- R54 · 36.0 Data Agent Confirmation ---
    if cloud and cloud.get("data_agent_enabled"):
        cells[54] = {
            "attribute": "Data Agent Confirmation",
            "B": _tag(
                "Cloud auto-upgrades Data Agent — confirm with customer before upgrade.",
                "Cloud Portal",
            ),
            "C": "Review",
        }
    else:
        cells[54] = {
            "attribute": "Data Agent Confirmation",
            "B": _tag("Data Agent not enabled, or cloud data unavailable.", "Cloud Portal"),
            "C": "N/A",
        }

    # =========================================================================
    # Section 8: FEATURE FLAGS & CONFIGURATION  (rows 56-57)
    # =========================================================================

    enabled_features = features.get("enabled_features", []) or features.get("features_enabled", [])
    disabled_features = features.get("disabled_features", []) or features.get("features_disabled", [])

    cells[56] = {
        "attribute": "Enabled Features",
        "B": _tag(", ".join(enabled_features) if enabled_features else "None reported", "CMC"),
        "C": "Done" if enabled_features else "Review",
    }
    cells[57] = {
        "attribute": "Disabled Features",
        "B": _tag(", ".join(disabled_features) if disabled_features else "None reported", "CMC"),
        "C": "Done",
    }

    return {**state, "cell_values": cells}


# ---------------------------------------------------------------------------
# Public entry: Write approved values to Excel, return as base64
# ---------------------------------------------------------------------------

def run_write_checklist_excel(
    cell_values_json: str = "",
    template_path: str = "",
    filename: str = "pre_upgrade_checklist_filled.xlsx",
    cmc_cluster_name: str = "",
) -> dict:
    """Write approved cell values into a copy of the Excel template.

    Produces two sheets: Summary (built dynamically) and Pre-Upgrade Checklist.

    Data-source resolution, in priority order:
      1. If `cmc_cluster_name` is provided and a cache file exists for it,
         use the cached payload. This is the normal path — avoids having
         the LLM ferry a large JSON blob between tool calls (which caused
         truncation / paraphrasing in earlier runs).
      2. Otherwise, parse `cell_values_json` (legacy / manual-override path).

    Args:
        cell_values_json: Fallback JSON string of {row_num: {"B": value, "C": status}}
            when no cached payload is available for the cluster.
        template_path: Path to the Excel template file.
        filename: Suggested filename for the download.
        cmc_cluster_name: Preferred — cluster key used to look up the cached
            payload written by generate_upgrade_readiness_report.

    Returns:
        dict with keys:
            - "type": "excel"
            - "filename": suggested download filename
            - "base64": base64-encoded .xlsx bytes
            - "summary": human-readable summary string
            - "source": "cache" or "inline_json" (where the payload came from)
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # --- Resolve the checklist payload ---
    raw = None
    source = "inline_json"
    if cmc_cluster_name:
        cached = load_checklist_cache(cmc_cluster_name)
        if cached is not None:
            raw = cached
            source = "cache"

    if raw is None:
        if not cell_values_json:
            raise ValueError(
                "No checklist payload available: pass cmc_cluster_name to use "
                "the server-side cache from generate_upgrade_readiness_report, "
                "or pass cell_values_json explicitly."
            )
        raw = json.loads(cell_values_json)

    # Don't mutate the cached dict for future reads
    raw = dict(raw)
    assessment = raw.pop("_summary", None)
    cell_values = {int(k): v for k, v in raw.items()}

    # Write into a temp file so we never need a caller-supplied output path
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        shutil.copy2(template_path, tmp_path)

        wb = load_workbook(tmp_path)

        # ------------------------------------------------------------------
        # Summary sheet (created from assessment data when available)
        # ------------------------------------------------------------------
        if assessment:
            _write_summary_sheet(wb, assessment, cell_values)

        # ------------------------------------------------------------------
        # Pre-Upgrade Checklist sheet
        # ------------------------------------------------------------------
        ws = wb["Pre-Upgrade Checklist"]

        status_fills = {
            "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "Done": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "Failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "WARNING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "Action Required": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "Pending": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            "N/A": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }

        # The new template pre-fills col A (#) and col B (Check Item name).
        # The writer fills col C (Details / Findings) and col D (Status).
        # The dict keys "B" and "C" are retained (legacy) but map to cols C and D.
        for row_num, cols in cell_values.items():
            # Col C — Details / Findings
            details_cell = ws.cell(row=row_num, column=3)
            details_cell.value = cols.get("B", "")
            details_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Col D — Status
            status = cols.get("C", "")
            status_cell = ws.cell(row=row_num, column=4)
            status_cell.value = status
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            fill = status_fills.get(status)
            if fill:
                status_cell.fill = fill

        # Fill the R9 meta row (merged A:D) with generation context from the assessment.
        if assessment:
            from datetime import datetime

            env = assessment.get("environment_summary", {}) or {}
            cluster = env.get("cluster_name") or env.get("cluster") or "—"
            rating = assessment.get("rating", "—")
            meta_text = (
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
                f"Cluster: {cluster} | "
                f"Verdict: {rating}"
            )
            ws.cell(row=9, column=1).value = meta_text

        wb.save(tmp_path)
        wb.close()

        with open(tmp_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("utf-8")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    filled_count = len(cell_values)
    not_impl_count = sum(1 for c in cell_values.values() if c.get("B") == "Not Implemented")

    summary = (
        f"# Checklist Generation Complete\n\n"
        f"- **Rows filled:** {filled_count}\n"
        f"- **Not Implemented rows:** {not_impl_count}\n"
        f"- **Sheets produced:** Summary, Pre-Upgrade Checklist\n"
        f"- **File:** {filename} (ready for download)\n"
    )

    return {
        "type": "excel",
        "filename": filename,
        "base64": encoded,
        "summary": summary,
        "source": source,
    }


# ---------------------------------------------------------------------------
# Summary sheet builder
# ---------------------------------------------------------------------------

def _write_summary_sheet(wb, assessment: dict, cell_values: dict):
    """Create a Summary sheet as the first tab with a readiness overview."""
    from collections import Counter
    from openpyxl.styles import Alignment, Font, PatternFill

    ss = wb.create_sheet("Summary", 0)
    ss.column_dimensions["A"].width = 60
    ss.column_dimensions["B"].width = 15

    bold = Font(bold=True)
    header_font = Font(bold=True, size=14)
    wrap = Alignment(wrap_text=True, vertical="top")

    rating = assessment.get("rating", "UNKNOWN")
    risk = assessment.get("risk_level", "UNKNOWN")
    rating_fills = {
        "READY": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "READY WITH CAVEATS": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "NOT READY": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    status_fills = {
        "Done": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "WARNING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Action Required": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Pending": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "N/A": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "Failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    row = 1

    # --- Header ---
    ss.cell(row=row, column=1, value="Upgrade Readiness Summary").font = header_font
    row += 1
    from_v = assessment.get("from_version", "?")
    to_v = assessment.get("to_version", "?")
    ss.cell(row=row, column=1, value=f"Upgrade: {from_v} \u2192 {to_v}").font = bold
    row += 1
    verdict_cell = ss.cell(row=row, column=1, value=f"Verdict: {rating} ({risk} RISK)")
    verdict_cell.font = Font(bold=True, size=12)
    verdict_cell.fill = rating_fills.get(rating, PatternFill())
    row += 1
    detail = assessment.get("rating_detail", "")
    if detail:
        ss.cell(row=row, column=1, value=detail).alignment = wrap
    row += 2

    # --- Status counts ---
    ss.cell(row=row, column=1, value="Status").font = bold
    ss.cell(row=row, column=2, value="Count").font = bold
    row += 1
    statuses = [v.get("C", "") for v in cell_values.values()]
    counts = Counter(statuses)
    for label in ["Done", "PASS", "Review", "WARNING", "Action Required", "Pending", "N/A", "Failed", "FAIL"]:
        if counts.get(label, 0) > 0:
            c_a = ss.cell(row=row, column=1, value=label)
            c_b = ss.cell(row=row, column=2, value=counts[label])
            fill = status_fills.get(label)
            if fill:
                c_a.fill = fill
            row += 1
    row += 1

    # --- Blockers ---
    blockers = assessment.get("blockers", [])
    ss.cell(row=row, column=1, value="Blockers").font = bold
    ss.cell(row=row, column=2, value=len(blockers)).font = bold
    if blockers:
        ss.cell(row=row, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    row += 1
    for b in blockers:
        ss.cell(row=row, column=1, value=f"- {b}").alignment = wrap
        row += 1
    row += 1

    # --- Warnings ---
    warnings = assessment.get("warnings", [])
    ss.cell(row=row, column=1, value="Warnings").font = bold
    ss.cell(row=row, column=2, value=len(warnings)).font = bold
    if warnings:
        ss.cell(row=row, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    row += 1
    for w in warnings:
        ss.cell(row=row, column=1, value=f"- {w}").alignment = wrap
        row += 1
    row += 1

    # --- Disclaimers (auto-generated from data gaps + known limitations) ---
    ss.cell(row=row, column=1, value="Disclaimers").font = bold
    row += 1
    data_gaps = assessment.get("data_gaps", [])
    for gap in data_gaps:
        ss.cell(row=row, column=1, value=f"- {gap}").alignment = wrap
        row += 1
    # Cloud-specific disclaimers
    env = assessment.get("environment_summary", {})
    if env.get("deployment", "").lower() == "cloud":
        ss.cell(row=row, column=1, value="- Memory values are allocated sizes, not actual machine memory").alignment = wrap
        row += 1
        ss.cell(row=row, column=1, value="- Per-pod disk utilization not available (API limited)").alignment = wrap
        row += 1
    if not data_gaps and not (env.get("deployment", "").lower() == "cloud"):
        ss.cell(row=row, column=1, value="- None").alignment = wrap
        row += 1
    row += 1

    # --- Need to Plan (placeholders) ---
    ss.cell(row=row, column=1, value="Need to Plan").font = bold
    row += 1
    for placeholder in [
        "- Access to machine",
        "- DB admin availability",
        "- Confirm existing Jar(s) with CSM/Release Management team",
    ]:
        ss.cell(row=row, column=1, value=placeholder).alignment = wrap
        row += 1
    row += 1

    # --- Down time / Approvals (placeholders) ---
    ss.cell(row=row, column=1, value="Down Time").font = bold
    row += 2
    ss.cell(row=row, column=1, value="Manager's Approval").font = bold
    row += 2
    ss.cell(row=row, column=1, value="Customer's Approval").font = bold
